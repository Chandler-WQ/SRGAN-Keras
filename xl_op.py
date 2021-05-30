from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from numpy import array


def insert_metrics(path, model_name, niqes, ssims, psnrs):
    wb = load_workbook(path)
    my_sheet = wb.worksheets[0]
    max_col = my_sheet.max_column
    insert_col_data(niqes, my_sheet, max_col + 1, model_name, "niqe")
    insert_col_data(ssims, my_sheet, max_col + 2, model_name, "ssim")
    insert_col_data(psnrs, my_sheet, max_col + 3, model_name, "psnr")
    wb.save(path)
    return


def insert_metrics_lpips(path, model_name, lpipss):
    wb = load_workbook(path)
    my_sheet = wb.worksheets[0]
    max_col = my_sheet.max_column
    insert_col_data(lpipss, my_sheet, max_col + 1, model_name, "lpips")
    wb.save(path)
    return


def insert_col_data(data, sheet, col, model_name, data_name):
    sheet[get_column_letter(col) + str(1)].value = model_name
    sheet[get_column_letter(col) + str(2)].value = data_name
    for i in range(len(data)):
        sheet[get_column_letter(col) + str(3 + i)] = data[i]

def read_data(path,col):
    wb = load_workbook(path)
    my_sheet= wb.worksheets[0]
    colum = my_sheet[col]
    data = []
    for x in range(len(colum)):
        data.append(colum[x].value)
    return data


